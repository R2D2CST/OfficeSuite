# Python native libraries
import os

# Third party libraries
import openpyxl
from openpyxl.utils import get_column_letter

# Self build libraries

from Func.AbstractDocument import AbstractDocument


class Excel(AbstractDocument):
    """
    Class reads a Excel file type or builds a new Excel file if does not exists
    Args:
        > filePath (str) : file path where the workbook exists
    Returns: None
    Attributes:
        > filePath (str) : file path where the workbook exists
        > workbook (workbook) : file object class
        > sheets (list[str]) : list of sheets contained in the workbook
        > workbookData (list[list[list[any]]]): 3D matrix (sheet, row, column)
    """

    def __init__(self, filePath: str):
        """
        We build the class starting object attributes.
            Args:
                > filePath (str) : file path where the workbook exists
            Returns: None
            Attributes:
                > filePath (str) : file path where the workbook exists
                > workbook (workbook) : file object class
                > sheets (list[str]) : list of sheets contained in the workbook
                > workbookData (list[list[list[any]]]): 3D matrix (sheet, row, column)
            Raises: None
        """

        # Class main attribute file path location
        self.filePath = filePath

        # We validate the given path corresponds with an .xlsx file type
        self.__validateFiletype()

        # We generate our first workbook attributes
        try:
            self.__readFile()

        # If file does not exists we build a new one
        except FileNotFoundError:
            self.__createNewFile()
            self.__readFile()

        pass

    def __validateFiletype(self) -> bool:
        """
        Method is private not intended to use outside this class type.
        Raises:
            f: Given file path: {self.filePath} Not a valid excel file type.

        Returns:
            bool: True if given path is a valid file path and an excel filetype, False otherwise.
        """
        if ".xlsx" not in self.filePath:
            raise f"Given file path: {self.filePath}\n Not a valid excel file type."

        else:
            return True

    def __readFile(self) -> any:
        """
        Method reads a Excel file and stores as object attributes
            Args: None
            Returns: None
            Attributes:
                > workbook (workbook) : file object class
                > sheets (list[str]) : list of sheets contained in the workbook
                > workbookData (list[list[list[any]]]): 3D matrix (sheet, row, column).
            Raises: None
        """
        self.workbook = openpyxl.load_workbook(filename=self.filePath)
        self.sheets = self.workbook.get_sheet_names()

        self.workbookData = list()

        # 3D matrix to store sheet data.
        self.workbookData = []

        for sheet in self.sheets:

            currentSheet = self.workbook[sheet]
            sheetMatrix = []  # 2D matrix for the current sheet.

            # Iterate through rows, returning only values.
            for row in currentSheet.iter_rows(values_only=True):

                # Append the row as a list.
                sheetMatrix.append(list(row))

            # Add the 2D matrix to the 3D list.
            self.workbookData.append(sheetMatrix)

        pass

    def __createNewFile(self) -> None:
        """
        Method builds a new empty excel file type.
            Args: None
            Returns: None
            Raises: None
        """
        workbook = openpyxl.Workbook()
        workbook.save(filename=self.filePath)
        del workbook
        pass

    def printWorkbookData(self):
        for sheetIndex, sheet in enumerate(self.workbookData):
            print(f"Sheet: {self.sheets[sheetIndex]}\nContent:\n")
            for row in sheet:
                print(row)
        pass

    pass
